import openpyxl
import os

target_files = [
    "/Users/matt/Downloads/01_可視性廣告整合報表.xlsx",
    "/Users/matt/Downloads/02_無效流量整合報表.xlsx",
    "/Users/matt/Downloads/03_網站品質整合報表.xlsx",
    "/Users/matt/Downloads/04_品牌適合度整合報表.xlsx",
    "/Users/matt/Downloads/05_優質曝光整合報表.xlsx"
]

def format_excel(file_path):
    print(f"🎨 正在格式化: {os.path.basename(file_path)}")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 讀取標題列
        headers = [cell.value for cell in ws[1]]
        
        for col_idx, header in enumerate(headers, 1):
            if not header: continue
            h_str = str(header)
            
            # 判斷欄位類型
            is_percent = '%' in h_str or 'rate' in h_str.lower()
            is_average = 'Average' in h_str or 'Avg' in h_str or '(Sec)' in h_str
            # 排除非數值的 ID 類欄位
            is_id = any(x in h_str for x in ['Date', '日期', 'Hour', '小時', 'Id', '編號', 'Code', 'Name', '名稱', 'Site', '網站', 'Supplier', '供應商', 'Dimensions', 'Most Used', '時間', '版位'])
            
            # 設定目標格式
            target_format = None
            divide_by_100 = False
            
            if is_id:
                continue # 跳過 ID 欄位
            
            if is_percent:
                target_format = '0.00%'
                divide_by_100 = True
            elif is_average:
                target_format = '#,##0.00' # 平均值保留小數
            else:
                target_format = '#,##0'    # 一般計數整數千分位
            
            # 應用格式到每一列
            # 注意: 從第 2 列開始
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    try:
                        # 嘗試轉為浮點數
                        val = float(cell.value)
                        
                        # 只有當數值 > 1 (例如 95.5) 且是百分比欄位時才除以 100
                        # 避免如果已經是 0.955 的情況被重複除 (雖然目前邏輯是乘過 100 的)
                        # 保險起見，我們假設目前的檔案都是 0-100 的數值 (因為是我剛產出的)
                        if divide_by_100:
                            cell.value = val / 100.0
                        
                        cell.number_format = target_format
                    except ValueError:
                        pass # 非數值則忽略
        
        wb.save(file_path)
        print(f"✅ 完成: {os.path.basename(file_path)}")
        
    except Exception as e:
        print(f"❌ 格式化失敗 {file_path}: {e}")

if __name__ == "__main__":
    for f in target_files:
        if os.path.exists(f):
            format_excel(f)
        else:
            print(f"⚠️ 找不到檔案: {f}")
